import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.drawing.image import Image

import datetime as dt
import pandas as pd
from Setup_info import bt_color, tt_color, rb_color, sb_color

# OPTIONAL PARAMETERS:
mark_empty_orders = True


# FETCHING DATES:
now = dt.datetime
present_date = now.today().strftime("%d-%m-%Y")

book_address = f'Logistics_{present_date}.xlsx'

workbook_name = book_address

alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']

# BORDER PARAMETERS
table_structure_border = Border(top=Side(border_style='thin'), bottom=Side(border_style='thin'),
                                right=Side(border_style='thin'), left=Side(border_style='thin'))
data_border = Border(top=Side(border_style='thin'), bottom=Side(border_style='thin'))
last_data_border = Border(top=Side(border_style='thin'), bottom=Side(border_style='thin'),
                          right=Side(border_style='thin'))

# BACKGROUND PARAMETERS
base_table_color = PatternFill(patternType='solid', fgColor=bt_color)
top_table_color = PatternFill(patternType='solid', fgColor=tt_color)
red_box_color = PatternFill(patternType='solid', fgColor=rb_color)
small_box_color = PatternFill(patternType='solid', fgColor=sb_color)
empty_order_style = PatternFill(patternType='lightGrid')

# FONT PARAMETERS
red_box_font = Font(size=14, color='890F0D')

# ALIGNMENT PARAMETERS (CENTER ONLY)
center_align = Alignment(horizontal='center', vertical='center')

# RAW TABLE STRUCTURE PARAMETERS
table_structure_columns = ['A', 'B']
table_structure_rows = [1]


def apply_style(sheet_name, table_height, table_width, red_box_index, small_box_index):
    wb = openpyxl.load_workbook(workbook_name)

    # DEFINING SHEET AND WORKING AREA
    ws = wb[sheet_name]

    # TABLE STRUCTURE AREA
    table_top_cells = []
    for i in range(table_width - 1):
        table_top_cells.append(str(alphabet[i]) + '1')

    table_structure_cells = []
    for column in table_structure_columns:
        for i in range(1, table_height + 1):
            table_structure_cells.append(str(column) + str(i))

    table_structure_cells.extend(table_top_cells)

    # DATA STRUCTURE
    data_columns = []
    data_cells = []
    for i in range(2, table_width - 1):
        data_columns.append(str(alphabet[i]))

    for column in data_columns:
        for i in range(2, table_height + 1):
            data_cells.append(str(column) + str(i))

    last_data_list = []
    last_data_column = data_columns[-1]
    for i in range(2, table_height + 1):
        last_data_list.append(str(last_data_column) + str(i))

    # BORDERS:

    # DATA:
    for i in data_cells:
        ws[i].border = data_border

    for i in last_data_list:
        ws[i].border = last_data_border

    # TABLE
    for i in table_structure_cells:
        ws[i].border = table_structure_border


    # STYLE:
    for i in table_structure_cells:
        ws[i].fill = base_table_color

    for i in table_top_cells:
        ws[i].fill = top_table_color

    for i in red_box_index:
        ws['B' + str(i)].fill = red_box_color
        ws['B' + str(i)].font = red_box_font

    for i in small_box_index:
        ws['B' + str(i)].fill = small_box_color


    # ALIGNMENT
    for i in table_structure_cells:
        ws[i].alignment = center_align

    for i in data_cells:
        ws[i].alignment = center_align


    # ADDING GRAY SHADING FOR EMPTY ORDERS
    gray_workbook_df = pd.read_excel(book_address, sheet_name=sheet_name, engine="openpyxl")
    gray_workbook_df.drop(columns=['Orders ⬇️', 'Box'], inplace=True)

    gray_rows = []
    gray_cells = []

    for index, row in gray_workbook_df.iterrows():
        empty_row = True
        for i in row.items():
            if isinstance(i[1], int):
                empty_row = False

        if empty_row:
            gray_rows.append(int(index) + 2)

    data_columns.append('A')
    data_columns.append('B')

    for column in data_columns:
        for row_number in gray_rows:
            gray_cells.append(str(column) + str(row_number))

    if mark_empty_orders:
        for i in gray_cells:
            ws[i].fill = empty_order_style


    # INSERT HEADER
    ws.insert_cols(1)

    ws['A1'] = sheet_name
    ws['A1'].alignment = center_align
    ws['A1'].fill = top_table_color
    ws['A1'].font = Font(bold=True)

    # ADDING IMAGE
    img = Image(f'Images/{sheet_name}.jpg')
    img.width = 210
    img.height = 210
    ws.add_image(img, 'A2')


    # ADJUST WIDTH
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 2
    ws.column_dimensions['A'].width = 30

    wb.save(workbook_name)
